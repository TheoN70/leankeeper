"""Generate the results.xlsx template for RAG evaluation."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ──────────────────────────────────────────
# Sheet 1: Raw Data
# ──────────────────────────────────────────

ws = wb.active
ws.title = "Raw Data"

CATEGORIES = ["Naming", "Generality", "Style", "API", "Attributes"]
MAX_ROWS = 50  # Support up to 50 PRs

# Styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
actual_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
rag_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center = Alignment(horizontal="center", vertical="center")

# Headers row 1: group headers
ws.merge_cells("A1:B1")
ws["A1"] = "PR Info"
ws["A1"].font = header_font
ws["A1"].fill = header_fill

ws.merge_cells("C1:G1")
ws["C1"] = "Actual (reviewer flagged: 1=yes, 0=no)"
ws["C1"].font = header_font
ws["C1"].fill = actual_fill
ws["C1"].alignment = center

ws.merge_cells("H1:L1")
ws["H1"] = "RAG (system flagged: 1=yes, 0=no)"
ws["H1"].font = header_font
ws["H1"].fill = rag_fill
ws["H1"].alignment = center

# Headers row 2: column names
col_headers = ["PR #", "Title"] + CATEGORIES + CATEGORIES
for i, h in enumerate(col_headers, 1):
    cell = ws.cell(row=2, column=i, value=h)
    cell.font = header_font
    cell.border = thin_border
    cell.alignment = center
    if i <= 2:
        cell.fill = header_fill
    elif i <= 7:
        cell.fill = actual_fill
    else:
        cell.fill = rag_fill

# Column widths
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 50
for col in range(3, 13):
    ws.column_dimensions[get_column_letter(col)].width = 14

# Data rows (empty, to be filled by agent)
for row in range(3, 3 + MAX_ROWS):
    for col in range(1, 13):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if col >= 3:
            cell.alignment = center

# Instructions row at bottom
instr_row = 3 + MAX_ROWS + 1
ws.cell(row=instr_row, column=1, value="Instructions:").font = Font(bold=True, italic=True)
ws.cell(row=instr_row + 1, column=1, value="Fill 1 if the category was flagged, 0 if not. One row per PR.")
ws.cell(row=instr_row + 2, column=1, value="'Actual' = what the real Mathlib reviewer flagged. 'RAG' = what LeanKeeper flagged.")
ws.cell(row=instr_row + 3, column=1, value="Categories: Naming (convention violations), Generality (over-specialized hypotheses),")
ws.cell(row=instr_row + 4, column=1, value="Style (proof tactics, formatting), API (missing lemmas/attributes), Attributes (@[simp], @[ext], etc.)")

# ──────────────────────────────────────────
# Sheet 2: Summary (auto-computed)
# ──────────────────────────────────────────

ws2 = wb.create_sheet("Summary")

title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=12)
metric_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

ws2["A1"] = "RAG Evaluation Summary"
ws2["A1"].font = title_font

# Data range for formulas
data_start = 3
data_end = 3 + MAX_ROWS - 1

# Per-category metrics
ws2["A3"] = "Per-Category Metrics"
ws2["A3"].font = section_font

headers = ["Category", "TP", "FP", "FN", "Precision", "Recall", "F1-Score"]
for i, h in enumerate(headers, 1):
    cell = ws2.cell(row=4, column=i, value=h)
    cell.font = header_font
    cell.fill = metric_fill
    cell.border = thin_border
    cell.alignment = center

# Actual columns: C=Naming(3), D=Generality(4), E=Style(5), F=API(6), G=Attributes(7)
# RAG columns:    H=Naming(8), I=Generality(9), J=Style(10), K=API(11), L=Attributes(12)

for idx, cat in enumerate(CATEGORIES):
    row = 5 + idx
    actual_col = get_column_letter(3 + idx)  # C, D, E, F, G
    rag_col = get_column_letter(8 + idx)     # H, I, J, K, L

    # Category name
    ws2.cell(row=row, column=1, value=cat).border = thin_border

    # TP: SUMPRODUCT where both actual=1 AND rag=1
    tp_formula = f"=SUMPRODUCT(('Raw Data'!{actual_col}{data_start}:{actual_col}{data_end}=1)*('Raw Data'!{rag_col}{data_start}:{rag_col}{data_end}=1))"
    ws2.cell(row=row, column=2, value=tp_formula).border = thin_border

    # FP: SUMPRODUCT where actual=0 AND rag=1
    fp_formula = f"=SUMPRODUCT(('Raw Data'!{actual_col}{data_start}:{actual_col}{data_end}=0)*('Raw Data'!{rag_col}{data_start}:{rag_col}{data_end}=1))"
    ws2.cell(row=row, column=3, value=fp_formula).border = thin_border

    # FN: SUMPRODUCT where actual=1 AND rag=0
    fn_formula = f"=SUMPRODUCT(('Raw Data'!{actual_col}{data_start}:{actual_col}{data_end}=1)*('Raw Data'!{rag_col}{data_start}:{rag_col}{data_end}=0))"
    ws2.cell(row=row, column=4, value=fn_formula).border = thin_border

    # Precision: TP / (TP + FP), handle division by zero
    tp_cell = f"B{row}"
    fp_cell = f"C{row}"
    fn_cell = f"D{row}"
    ws2.cell(row=row, column=5, value=f'=IF(({tp_cell}+{fp_cell})=0,"N/A",{tp_cell}/({tp_cell}+{fp_cell}))').border = thin_border

    # Recall: TP / (TP + FN)
    ws2.cell(row=row, column=6, value=f'=IF(({tp_cell}+{fn_cell})=0,"N/A",{tp_cell}/({tp_cell}+{fn_cell}))').border = thin_border

    # F1: 2 * P * R / (P + R)
    p_cell = f"E{row}"
    r_cell = f"F{row}"
    ws2.cell(row=row, column=7, value=f'=IF(OR({p_cell}="N/A",{r_cell}="N/A",({p_cell}+{r_cell})=0),"N/A",2*{p_cell}*{r_cell}/({p_cell}+{r_cell}))').border = thin_border

    # Format percentages
    for col in [5, 6, 7]:
        ws2.cell(row=row, column=col).number_format = '0.0%'
        ws2.cell(row=row, column=col).alignment = center

    for col in [2, 3, 4]:
        ws2.cell(row=row, column=col).alignment = center

# Overall metrics
overall_row = 5 + len(CATEGORIES) + 1
ws2.cell(row=overall_row, column=1, value="OVERALL").font = Font(bold=True)
ws2.cell(row=overall_row, column=1).border = thin_border

# Overall TP/FP/FN: sum of all categories
for col_idx, label in [(2, "B"), (3, "C"), (4, "D")]:
    formula = f"=SUM({label}5:{label}{5 + len(CATEGORIES) - 1})"
    cell = ws2.cell(row=overall_row, column=col_idx, value=formula)
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.alignment = center

# Overall Precision, Recall, F1
tp_cell = f"B{overall_row}"
fp_cell = f"C{overall_row}"
fn_cell = f"D{overall_row}"
ws2.cell(row=overall_row, column=5, value=f'=IF(({tp_cell}+{fp_cell})=0,"N/A",{tp_cell}/({tp_cell}+{fp_cell}))').border = thin_border
ws2.cell(row=overall_row, column=6, value=f'=IF(({tp_cell}+{fn_cell})=0,"N/A",{tp_cell}/({tp_cell}+{fn_cell}))').border = thin_border
p_cell = f"E{overall_row}"
r_cell = f"F{overall_row}"
ws2.cell(row=overall_row, column=7, value=f'=IF(OR({p_cell}="N/A",{r_cell}="N/A",({p_cell}+{r_cell})=0),"N/A",2*{p_cell}*{r_cell}/({p_cell}+{r_cell}))').border = thin_border

for col in [5, 6, 7]:
    ws2.cell(row=overall_row, column=col).number_format = '0.0%'
    ws2.cell(row=overall_row, column=col).alignment = center
    ws2.cell(row=overall_row, column=col).font = Font(bold=True)

# Column widths
ws2.column_dimensions["A"].width = 14
for col in range(2, 8):
    ws2.column_dimensions[get_column_letter(col)].width = 14

# Count row
count_row = overall_row + 2
ws2.cell(row=count_row, column=1, value="PRs evaluated:").font = Font(bold=True)
ws2.cell(row=count_row, column=2, value=f"=COUNTA('Raw Data'!A{data_start}:A{data_end})")

# Save
output = "results/results.xlsx"
import os
os.makedirs("results", exist_ok=True)
wb.save(output)
print(f"Template saved to {output}")
